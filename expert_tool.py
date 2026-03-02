# ==============================================================
# AHP EXPERT INPUT TOOL — Flask Web App
# Launched automatically by newcr.py
# Can also run standalone: python3 expert_tool.py
# ==============================================================

import os, sys, json, argparse, threading, webbrowser
import pandas as pd
from flask import Flask, request, jsonify, render_template_string
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Parse args ─────────────────────────────────────────────────
parser = argparse.ArgumentParser(add_help=False)
parser.add_argument("--mode",     default="standalone")
parser.add_argument("--sentinel", default=None)   # explicit path from newcr.py
args, _ = parser.parse_known_args()

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE  = os.path.join(BASE_DIR, "ahp_input.xlsx")

# Sentinel: use explicit path if provided by newcr.py, else default
SENTINEL = args.sentinel if args.sentinel else os.path.join(BASE_DIR, ".expert_saved")

app = Flask(__name__)

# ── Helpers ────────────────────────────────────────────────────
def bd():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(c):
    return PatternFill("solid", start_color=c)

def read_current_criteria():
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="Criteria_Config", header=2)
        df.columns = df.columns.str.strip()
        df = df.dropna(subset=["Criteria Name"])
        df = df[~df["Criteria Name"].astype(str).str.startswith("⚠")]
        df = df[df["Criteria Name"].astype(str).str.strip() != ""]
        return [{"name": str(r["Criteria Name"]).strip(),
                 "type": str(r["Type (Benefit/Cost)"]).strip().lower()}
                for _, r in df.iterrows()]
    except Exception:
        return [
            {"name":"TravelTime","type":"cost"},
            {"name":"Comfort",   "type":"benefit"},
            {"name":"Cost",      "type":"cost"},
            {"name":"Frequency", "type":"benefit"},
        ]

def write_to_excel(criteria_names, benefit_list, matrix_dict):
    wb = load_workbook(EXCEL_FILE)
    DARK="1F4E79"; MID="2E75B6"; LB="D9E2F3"; AB="EBF3FB"; WHITE="FFFFFF"
    n = len(criteria_names)

    def write_matrix_sheet(ws):
        # Row 1 = headers (no decorative title row)
        ws.cell(1,1,"Criteria").font = Font(name="Arial",bold=True,color=WHITE,size=10)
        ws.cell(1,1).fill = fill(DARK)
        ws.cell(1,1).alignment = Alignment(horizontal="center",vertical="center")
        ws.cell(1,1).border = bd()
        for ci,crit in enumerate(criteria_names,2):
            c = ws.cell(1,ci,crit)
            c.font = Font(name="Arial",bold=True,color=WHITE,size=10)
            c.fill = fill(DARK); c.alignment = Alignment(horizontal="center"); c.border = bd()
        for ri,rc in enumerate(criteria_names):
            row = ri+2; bg = LB if ri%2==0 else AB
            c0 = ws.cell(row,1,rc)
            c0.font = Font(name="Arial",bold=True,color=DARK,size=10)
            c0.fill = fill(bg); c0.alignment = Alignment(horizontal="left",vertical="center"); c0.border = bd()
            for ci,_ in enumerate(criteria_names):
                val = matrix_dict[rc][ci]
                cell = ws.cell(row,ci+2,round(val,4))
                cell.font = Font(name="Arial",size=10)
                cell.fill = fill(bg); cell.alignment = Alignment(horizontal="center",vertical="center")
                cell.border = bd(); cell.number_format = "0.0000"
        for i in range(n+1):
            ws.column_dimensions[get_column_letter(i+1)].width = 16

    if "Criteria" in wb.sheetnames: del wb["Criteria"]
    ws_c = wb.create_sheet("Criteria", 0)
    write_matrix_sheet(ws_c)

    if "Expert_Input" in wb.sheetnames: del wb["Expert_Input"]
    pos = wb.sheetnames.index("Criteria_Config")+1 if "Criteria_Config" in wb.sheetnames else len(wb.sheetnames)
    ws_e = wb.create_sheet("Expert_Input", pos)
    write_matrix_sheet(ws_e)

    if "Criteria_Config" in wb.sheetnames:
        ws_cfg = wb["Criteria_Config"]
        for row in ws_cfg.iter_rows(min_row=4, max_row=ws_cfg.max_row):
            cname = str(row[0].value).strip() if row[0].value else ""
            if cname in criteria_names:
                ctype = "Benefit" if cname in benefit_list else "Cost"
                row[1].value = ctype
                row[1].font = Font(name="Arial",size=10,bold=True,
                                   color="375623" if ctype=="Benefit" else "C00000")
                row[1].alignment = Alignment(horizontal="center",vertical="center")

    order = ["Criteria","Alternatives_Data","Criteria_Config","Expert_Input","Results"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
    wb.save(EXCEL_FILE)

    # Write sentinel so newcr.py knows to continue
    with open(SENTINEL, "w") as f:
        f.write("saved")
    print(f"[expert_tool] ✔ Excel saved. Sentinel written → {SENTINEL}")

# ── Flask routes ───────────────────────────────────────────────
@app.route("/")
def index():
    criteria = read_current_criteria()
    return render_template_string(HTML_PAGE, criteria_json=json.dumps(criteria))

@app.route("/submit", methods=["POST"])
def submit():
    data = request.get_json()
    criteria_names = [c["name"] for c in data["criteria"]]
    benefit_list   = [c["name"] for c in data["criteria"] if c["type"]=="benefit"]
    matrix_dict    = data["matrix"]
    try:
        write_to_excel(criteria_names, benefit_list, matrix_dict)
        return jsonify({"status":"ok"})
    except Exception as e:
        return jsonify({"status":"error","message":str(e)}), 500

# ── HTML (same polished UI) ────────────────────────────────────
HTML_PAGE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/><meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>AHP Expert Input — Bus Allocation</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display&family=DM+Mono:wght@400;500&family=Outfit:wght@300;400;500;600;700&display=swap" rel="stylesheet"/>
<style>
:root{--bg:#0d1117;--surface:#161b22;--card:#1c2230;--border:#2d3748;--accent:#3b82f6;--accent2:#06b6d4;--gold:#f59e0b;--green:#10b981;--red:#ef4444;--text:#e2e8f0;--muted:#8892a4;--mono:'DM Mono',monospace;--serif:'DM Serif Display',serif;--sans:'Outfit',sans-serif}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--text);font-family:var(--sans);min-height:100vh}
body::before{content:'';position:fixed;inset:0;background-image:linear-gradient(rgba(59,130,246,.04) 1px,transparent 1px),linear-gradient(90deg,rgba(59,130,246,.04) 1px,transparent 1px);background-size:48px 48px;pointer-events:none;z-index:0}
.wrap{position:relative;z-index:1;max-width:960px;margin:0 auto;padding:2rem 1.5rem 4rem}
.hero{text-align:center;padding:2rem 0 1.8rem}
.badge{display:inline-block;background:linear-gradient(135deg,#1e3a5f,#1e4d6b);border:1px solid rgba(59,130,246,.35);border-radius:999px;padding:.28rem 1rem;font-size:.7rem;font-weight:600;letter-spacing:.12em;text-transform:uppercase;color:var(--accent2);margin-bottom:.9rem}
.hero h1{font-family:var(--serif);font-size:clamp(1.8rem,4vw,2.7rem);background:linear-gradient(135deg,#e2e8f0 30%,#60a5fa);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:.5rem}
.hero p{color:var(--muted);font-size:.88rem;max-width:500px;margin:0 auto;line-height:1.6}
.steps{display:flex;justify-content:center;margin:1.4rem 0;position:relative}
.steps::before{content:'';position:absolute;top:50%;left:10%;right:10%;height:1px;background:var(--border);transform:translateY(-50%);z-index:0}
.step{display:flex;flex-direction:column;align-items:center;gap:.3rem;position:relative;z-index:1;flex:1;max-width:110px}
.dot{width:32px;height:32px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:.78rem;font-weight:700;border:2px solid var(--border);background:var(--surface);color:var(--muted);transition:all .3s}
.step.active .dot{border-color:var(--accent);background:rgba(59,130,246,.15);color:var(--accent);box-shadow:0 0 12px rgba(59,130,246,.28)}
.step.done .dot{border-color:var(--green);background:rgba(16,185,129,.15);color:var(--green)}
.slbl{font-size:.62rem;color:var(--muted);text-align:center;font-weight:500;letter-spacing:.04em}
.step.active .slbl{color:var(--accent)}.step.done .slbl{color:var(--green)}
.panel{background:var(--card);border:1px solid var(--border);border-radius:14px;padding:1.8rem;margin-bottom:1.4rem;display:none;animation:fu .35s ease}
.panel.active{display:block}
@keyframes fu{from{opacity:0;transform:translateY(12px)}to{opacity:1;transform:translateY(0)}}
.pt{font-family:var(--serif);font-size:1.25rem;color:var(--text);margin-bottom:.28rem}
.ps{color:var(--muted);font-size:.82rem;margin-bottom:1.3rem;line-height:1.55}
.clist{display:flex;flex-direction:column;gap:.65rem}
.ci{display:flex;align-items:center;gap:.85rem;background:var(--surface);border:1px solid var(--border);border-radius:9px;padding:.7rem .9rem;transition:border-color .2s}
.ci:hover{border-color:rgba(59,130,246,.4)}
.cn{width:24px;height:24px;border-radius:5px;background:rgba(59,130,246,.12);border:1px solid rgba(59,130,246,.25);display:flex;align-items:center;justify-content:center;font-size:.72rem;font-weight:700;color:var(--accent);flex-shrink:0}
.cinp{flex:1;background:transparent;border:none;outline:none;color:var(--text);font-family:var(--sans);font-size:.9rem;font-weight:500}
.cinp::placeholder{color:var(--muted);font-weight:400}
.tb{padding:.2rem .6rem;border-radius:999px;font-size:.68rem;font-weight:700;letter-spacing:.06em;cursor:pointer;transition:all .2s;user-select:none;border:1px solid transparent;flex-shrink:0}
.tb.benefit{background:rgba(16,185,129,.15);border-color:rgba(16,185,129,.3);color:var(--green)}
.tb.cost{background:rgba(239,68,68,.12);border-color:rgba(239,68,68,.25);color:var(--red)}
.db{background:none;border:none;color:var(--muted);cursor:pointer;padding:.18rem;border-radius:4px;transition:color .2s;font-size:.9rem;flex-shrink:0}
.db:hover{color:var(--red)}
.ab{display:flex;align-items:center;justify-content:center;gap:.4rem;width:100%;padding:.6rem;border:1.5px dashed var(--border);border-radius:9px;background:transparent;color:var(--muted);font-family:var(--sans);font-size:.84rem;cursor:pointer;transition:all .2s;margin-top:.45rem}
.ab:hover{border-color:var(--accent);color:var(--accent)}
.cc{background:var(--surface);border:1px solid var(--border);border-radius:13px;padding:1.3rem;margin-bottom:1rem;transition:border-color .3s}
.cc:hover{border-color:rgba(59,130,246,.3)}
.vh{display:flex;align-items:center;justify-content:center;gap:.6rem;margin-bottom:.9rem;flex-wrap:wrap}
.vl{font-size:.93rem;font-weight:600;padding:.35rem .8rem;border:1px solid rgba(59,130,246,.2);border-radius:7px;background:rgba(59,130,246,.06);white-space:nowrap}
.vq{text-align:center;font-size:.82rem;color:var(--muted);margin-bottom:1rem;font-style:italic}
.mb{display:flex;justify-content:space-between;font-size:.64rem;color:var(--muted);margin-bottom:.4rem}
.sr{display:grid;grid-template-columns:repeat(17,1fr);gap:.2rem;margin-bottom:.65rem}
.sb{border-radius:6px;border:1.5px solid var(--border);background:var(--card);color:var(--muted);font-family:var(--mono);font-size:.72rem;font-weight:600;cursor:pointer;transition:all .15s;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:.28rem .05rem;gap:1px;aspect-ratio:.82}
.sb:hover{border-color:var(--accent);color:var(--text);transform:translateY(-2px)}
.sb.sl{border-color:var(--accent);background:rgba(59,130,246,.18);color:var(--accent);box-shadow:0 0 10px rgba(59,130,246,.18);transform:translateY(-2px)}
.sb.sr2{border-color:var(--accent2);background:rgba(6,182,212,.11);color:var(--accent2);box-shadow:0 0 10px rgba(6,182,212,.16);transform:translateY(-2px)}
.sb.se{border-color:var(--gold);background:rgba(245,158,11,.1);color:var(--gold);transform:translateY(-2px)}
.sn{font-size:.78rem;font-weight:700}
.st{font-size:.4rem;font-family:var(--sans);opacity:.75;text-align:center;line-height:1.2}
.eb{display:block;width:100%;padding:.42rem;border:1.5px dashed var(--border);background:transparent;border-radius:6px;color:var(--muted);font-family:var(--sans);font-size:.76rem;cursor:pointer;transition:all .2s;margin-top:.3rem}
.eb:hover{border-color:var(--gold);color:var(--gold)}
.ptxt{font-size:.74rem;color:var(--muted);text-align:right;margin-bottom:.32rem}
.pw{background:var(--border);border-radius:999px;height:4px;margin-bottom:1.3rem;overflow:hidden}
.pf{height:100%;background:linear-gradient(90deg,var(--accent),var(--accent2));border-radius:999px;transition:width .4s}
.mw{overflow-x:auto;margin-top:.7rem}
.mt{width:100%;border-collapse:collapse;font-size:.76rem}
.mt th,.mt td{border:1px solid var(--border);padding:.4rem .5rem;text-align:center;font-family:var(--mono)}
.mt th{background:rgba(59,130,246,.1);color:var(--accent);font-size:.68rem}
.mt td{color:var(--text)}.mt td.dg{color:var(--muted)}.mt td.ab2{color:var(--accent)}.mt td.bl{color:var(--accent2)}
.crb{display:flex;gap:.8rem;flex-wrap:wrap;margin-top:.9rem}
.crc{flex:1;min-width:120px;background:var(--surface);border:1px solid var(--border);border-radius:11px;padding:.85rem;text-align:center}
.crl{font-size:.68rem;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;margin-bottom:.32rem}
.crv{font-family:var(--mono);font-size:1.45rem;font-weight:600}
.crv.ok{color:var(--green)}.crv.warn{color:var(--gold)}.crv.fail{color:var(--red)}
.crs{font-size:.7rem;margin-top:.22rem}
.br{display:flex;gap:.7rem;justify-content:flex-end;margin-top:1.3rem;flex-wrap:wrap}
.btn{padding:.58rem 1.3rem;border-radius:9px;font-family:var(--sans);font-size:.86rem;font-weight:600;cursor:pointer;border:none;transition:all .2s;display:flex;align-items:center;gap:.38rem}
.bp{background:linear-gradient(135deg,#2563eb,#0891b2);color:#fff;box-shadow:0 4px 12px rgba(37,99,235,.26)}
.bp:hover{transform:translateY(-1px);box-shadow:0 6px 16px rgba(37,99,235,.36)}
.bs{background:var(--surface);color:var(--text);border:1px solid var(--border)}
.bs:hover{border-color:var(--accent);color:var(--accent)}
.bsu{background:linear-gradient(135deg,#059669,#0891b2);color:#fff;box-shadow:0 4px 12px rgba(5,150,105,.26)}
.bsu:hover{transform:translateY(-1px);box-shadow:0 6px 16px rgba(5,150,105,.36)}
.bsu:disabled{opacity:.5;cursor:not-allowed;transform:none}
.ib{background:rgba(59,130,246,.07);border:1px solid rgba(59,130,246,.2);border-radius:9px;padding:.85rem 1rem;font-size:.8rem;color:var(--muted);line-height:1.58;margin-bottom:1rem}
.ib strong{color:var(--accent2)}
.sov{display:none;position:fixed;inset:0;background:rgba(0,0,0,.78);backdrop-filter:blur(6px);z-index:100;align-items:center;justify-content:center}
.sov.show{display:flex}
.sbx{background:var(--card);border:1px solid rgba(16,185,129,.4);border-radius:18px;padding:2.8rem 2.4rem;text-align:center;max-width:440px;animation:fu .4s ease}
.si{font-size:3.2rem;margin-bottom:.9rem}
.stl{font-family:var(--serif);font-size:1.7rem;color:var(--green);margin-bottom:.65rem}
.smg{color:var(--muted);font-size:.86rem;line-height:1.65;margin-bottom:1.4rem}
.smg code{color:var(--accent2);font-family:var(--mono);font-size:.84rem}
</style>
</head>
<body>
<div class="wrap">
  <div class="hero">
    <div class="badge">AHP · Bus Allocation · Expert Input</div>
    <h1>Pairwise Comparison Tool</h1>
    <p>Rate how important each criterion is compared to others. Click <strong>Save to Excel</strong> — the analysis runs automatically.</p>
  </div>
  <div class="steps">
    <div class="step active" id="si1"><div class="dot">1</div><div class="slbl">Criteria</div></div>
    <div class="step" id="si2"><div class="dot">2</div><div class="slbl">Compare</div></div>
    <div class="step" id="si3"><div class="dot">3</div><div class="slbl">Save</div></div>
  </div>

  <div class="panel active" id="p1">
    <div class="pt">Define Criteria</div>
    <div class="ps">Pre-loaded from your Excel. Toggle <span style="color:var(--green);font-weight:600">BENEFIT</span>/<span style="color:var(--red);font-weight:600">COST</span> by clicking the badge.</div>
    <div class="clist" id="cl"></div>
    <button class="ab" onclick="addC()">＋ Add Criterion</button>
    <div class="br"><button class="btn bp" onclick="gS2()">Compare Pairs →</button></div>
  </div>

  <div class="panel" id="p2">
    <div class="pt">Pairwise Comparisons</div>
    <div class="ps">Click a number: <span style="color:var(--accent)">blue</span>=left preferred · <span style="color:var(--accent2)">cyan</span>=right preferred · <span style="color:var(--gold)">1</span>=equal.</div>
    <div class="ib"><strong>Saaty Scale:</strong> 1=Equal · 3=Moderate · 5=Strong · 7=Very Strong · 9=Extreme · 2,4,6,8=Intermediate</div>
    <div class="ptxt" id="ptxt">0/0</div>
    <div class="pw"><div class="pf" id="pf" style="width:0%"></div></div>
    <div id="cw"></div>
    <div class="br">
      <button class="btn bs" onclick="gS1()">← Back</button>
      <button class="btn bp" onclick="gS3()">Review →</button>
    </div>
  </div>

  <div class="panel" id="p3">
    <div class="pt">Review & Save</div>
    <div class="ps">Check your weights and CR. Click Save — your terminal will continue automatically.</div>
    <div id="crd"></div>
    <div style="margin-top:1.2rem">
      <div style="font-size:.76rem;color:var(--muted);margin-bottom:.4rem;font-weight:600;letter-spacing:.05em;text-transform:uppercase">Pairwise Matrix</div>
      <div class="mw" id="mp"></div>
    </div>
    <div class="br">
      <button class="btn bs" onclick="gS2()">← Edit</button>
      <button class="btn bsu" id="svbtn" onclick="saveXL()">💾 Save to Excel</button>
    </div>
  </div>
</div>

<div class="sov" id="sov">
  <div class="sbx">
    <div class="si">✅</div>
    <div class="stl">Saved!</div>
    <div class="smg"><code>ahp_input.xlsx</code> has been updated.<br/><br/>
      Your terminal is now running the AHP–WSM analysis.<br/>
      Check the terminal for the full bus allocation ranking.</div>
    <button class="btn bsu" style="margin:0 auto" onclick="document.getElementById('sov').classList.remove('show')">Close</button>
  </div>
</div>

<script>
let criteria={{ criteria_json | safe }}, M={};
function setStep(n){[1,2,3].forEach(s=>{document.getElementById('p'+s).classList.toggle('active',s===n);const d=document.getElementById('si'+s);d.classList.remove('active','done');if(s<n)d.classList.add('done');else if(s===n)d.classList.add('active');});}
function gS1(){setStep(1);renderC();}
function gS2(){criteria=criteria.filter(c=>c.name.trim()!=='');if(criteria.length<2){alert('Need ≥2 criteria.');return;}initM();renderCmp();setStep(2);}
function gS3(){const miss=getPairs().filter(([i,j])=>M[i][j]===null);if(miss.length){alert(`Answer ${miss.length} more comparison(s).`);return;}renderRes();setStep(3);}
function renderC(){const l=document.getElementById('cl');l.innerHTML='';criteria.forEach((c,i)=>{const d=document.createElement('div');d.className='ci';d.innerHTML=`<div class="cn">${i+1}</div><input class="cinp" value="${c.name}" placeholder="Name" oninput="criteria[${i}].name=this.value"/><span class="tb ${c.type}" onclick="tT(${i})">${c.type.toUpperCase()}</span><button class="db" onclick="rmC(${i})">✕</button>`;l.appendChild(d);});}
function addC(){criteria.push({name:'',type:'benefit'});renderC();document.querySelectorAll('.cinp')[criteria.length-1].focus();}
function rmC(i){if(criteria.length<=2){alert('Need ≥2.');return;}criteria.splice(i,1);renderC();}
function tT(i){criteria[i].type=criteria[i].type==='benefit'?'cost':'benefit';renderC();}
function getPairs(){const p=[];for(let i=0;i<criteria.length;i++)for(let j=i+1;j<criteria.length;j++)p.push([i,j]);return p;}
function initM(){criteria.forEach((_,i)=>{if(!M[i])M[i]={};criteria.forEach((_,j)=>{if(i!==j&&M[i][j]===undefined)M[i][j]=null;});});}
function pick(i,j,v){v=parseFloat(v);M[i][j]=v;M[j]=M[j]||{};M[j][i]=1/v;document.getElementById(`sr${i}-${j}`).innerHTML=bldS(i,j);updP();}
function bldS(i,j){const cur=M[i][j];let h='';
for(let v=9;v>=2;v--){const s=cur!==null&&Math.abs(cur-v)<.001?'sl':'';h+=`<button class="sb ${s}" onclick="pick(${i},${j},${v})"><span class="sn">${v}</span><span class="st" style="color:var(--accent)">${criteria[i].name.slice(0,4)}</span></button>`;}
const se=cur===1?'se':'';h+=`<button class="sb ${se}" onclick="pick(${i},${j},1)" style="border-color:rgba(245,158,11,.3)"><span class="sn" style="color:var(--gold)">1</span><span class="st" style="color:var(--gold)">Equal</span></button>`;
for(let v=2;v<=9;v++){const inv=1/v,s=cur!==null&&Math.abs(cur-inv)<.0001?'sr2':'';h+=`<button class="sb ${s}" onclick="pick(${i},${j},${inv})"><span class="sn">${v}</span><span class="st" style="color:var(--accent2)">${criteria[j].name.slice(0,4)}</span></button>`;}
return h;}
function renderCmp(){const w=document.getElementById('cw');w.innerHTML='';getPairs().forEach(([i,j])=>{const c=document.createElement('div');c.className='cc';c.innerHTML=`<div class="vh"><span class="vl" style="color:var(--accent)">${criteria[i].name}</span><span style="font-size:.7rem;font-weight:700;color:var(--muted)">VS</span><span class="vl" style="color:var(--accent2)">${criteria[j].name}</span></div><div class="vq">How important is <em>${criteria[i].name}</em> compared to <em>${criteria[j].name}</em>?</div><div class="mb"><span style="color:var(--accent)">← ${criteria[i].name} preferred</span><span style="color:var(--gold)">Equal</span><span style="color:var(--accent2)">${criteria[j].name} preferred →</span></div><div class="sr" id="sr${i}-${j}">${bldS(i,j)}</div><button class="eb" onclick="pick(${i},${j},1)">⟺ Equal</button>`;w.appendChild(c);});updP();}
function updP(){const pairs=getPairs(),done=pairs.filter(([i,j])=>M[i]&&M[i][j]!==null).length,tot=pairs.length;document.getElementById('ptxt').textContent=`${done} / ${tot} answered`;document.getElementById('pf').style.width=`${tot?done/tot*100:0}%`;}
function bldM(){const n=criteria.length;return Array.from({length:n},(_,i)=>Array.from({length:n},(_,j)=>i===j?1:(M[i]&&M[i][j]!==null?M[i][j]:1)));}
function calcCR(m){const n=m.length,cs=m[0].map((_,j)=>m.reduce((s,r)=>s+r[j],0)),norm=m.map(r=>r.map((v,j)=>v/cs[j])),w=norm.map(r=>r.reduce((s,v)=>s+v,0)/n),ws=m.map((r,i)=>r.reduce((s,v,j)=>s+v*w[j],0)),lm=ws.reduce((s,v,i)=>s+v/w[i],0)/n,CI=(lm-n)/(n-1),RI={1:0,2:0,3:.58,4:.9,5:1.12,6:1.24,7:1.32,8:1.41,9:1.45,10:1.49};return{w,CI,CR:RI[n]?CI/RI[n]:0};}
function renderRes(){const m=bldM(),{w,CI,CR}=calcCR(m),n=criteria.length,cc=CR<.1?'ok':CR<.15?'warn':'fail',cm=CR<.1?'✔ Consistent':CR<.15?'⚠ Borderline':'✘ Revise';document.getElementById('crd').innerHTML=`<div class="crb">${criteria.map((c,i)=>`<div class="crc"><div class="crl">${c.name}</div><div class="crv ok">${(w[i]*100).toFixed(1)}%</div><div class="crs" style="color:var(--muted)">Weight</div></div>`).join('')}<div class="crc"><div class="crl">Consistency Ratio</div><div class="crv ${cc}">${CR.toFixed(4)}</div><div class="crs">${cm}</div></div></div>`;
let t=`<table class="mt"><thead><tr><th></th>${criteria.map(c=>`<th>${c.name}</th>`).join('')}</tr></thead><tbody>`;m.forEach((r,i)=>{t+=`<tr><th style="text-align:left;color:var(--text)">${criteria[i].name}</th>${r.map((v,j)=>`<td class="${i===j?'dg':i<j?'ab2':'bl'}">${i===j?'1':v>=1?v.toFixed(2):'1/'+Math.round(1/v)}</td>`).join('')}</tr>`;});t+=`</tbody></table>`;document.getElementById('mp').innerHTML=t;}
function saveXL(){const btn=document.getElementById('svbtn');btn.disabled=true;btn.textContent='Saving…';const m=bldM(),md={};criteria.forEach((c,i)=>{md[c.name]=m[i];});
fetch('/submit',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({criteria,matrix:md})}).then(r=>r.json()).then(d=>{btn.disabled=false;btn.textContent='💾 Save to Excel';if(d.status==='ok')document.getElementById('sov').classList.add('show');else alert('Error: '+d.message);}).catch(e=>{btn.disabled=false;btn.textContent='💾 Save to Excel';alert('Error: '+e);});}
renderC();
</script>
</body>
</html>"""

# ── Launch ─────────────────────────────────────────────────────
if __name__ == "__main__":
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: '{EXCEL_FILE}' not found in {BASE_DIR}")
        sys.exit(1)

    # Clean old sentinel
    if os.path.exists(SENTINEL):
        os.remove(SENTINEL)

    port = 5050

    if args.mode == "standalone":
        # Opened directly by user — open browser ourselves
        threading.Timer(1.3, lambda: webbrowser.open(f"http://127.0.0.1:{port}")).start()
        print(f"\n  AHP Expert Tool  →  http://127.0.0.1:{port}")
        print(f"  Ctrl+C to stop\n")

    print(f"  [Sentinel path: {SENTINEL}]")
    app.run(port=port, debug=False, use_reloader=False)
