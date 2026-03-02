# ==============================================================
# HYBRID AHP – WEIGHTED SUM MODEL (WSM)
# Project: MULTI CRITERIA DECISION MAKING TOOL FOR MASS TRANSPORTATION USING AHP
# DONE BY : Raja Gopal R , SASVAT BAALAN A , RAHMAN SHERIEF A
# ==============================================================

import numpy as np
import pandas as pd
import os, sys, subprocess, time, webbrowser
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE  = os.path.join(BASE_DIR, "ahp_input.xlsx")
EXPERT_TOOL = os.path.join(BASE_DIR, "expert_tool.py")
SENTINEL    = os.path.join(BASE_DIR, ".expert_saved")

# ══════════════════════════════════════════════════════════════
# 1. READ CRITERIA MATRIX
# ══════════════════════════════════════════════════════════════
def read_criteria_matrix(excel_file):
    """Read criteria matrix — auto-detects header row position."""
    from openpyxl import load_workbook as _lw
    _ws = _lw(excel_file)["Criteria"]
    header_row = 2  # default: row 3 in Excel (0-indexed=2)
    for ri, row in enumerate(_ws.iter_rows(min_row=1, max_row=4)):
        strs = [c.value for c in row if isinstance(c.value, str)]
        if len(strs) >= 2:
            header_row = ri
            break
    df = pd.read_excel(excel_file, sheet_name="Criteria", index_col=0, header=header_row)
    df.columns = [str(c).split(" (")[0].strip() for c in df.columns]
    df.index   = [str(i).split(" (")[0].strip() for i in df.index]
    df = df.dropna(how="all").dropna(axis=1, how="all")
    df = df.apply(pd.to_numeric, errors="coerce").dropna()
    return df

# ══════════════════════════════════════════════════════════════
# 2. READ ALTERNATIVES DATA
#    Sheet has: Bus(index) | Route | Type | TravelTime | Comfort | Cost | Frequency
#    Row 2 is a units sub-header — skipped automatically.
# ══════════════════════════════════════════════════════════════
def read_alternatives(excel_file, criteria_names):
    # skiprows=[1] skips the units sub-header (Excel row 2)
    df = pd.read_excel(excel_file, sheet_name="Alternatives_Data",
                       index_col=0, header=0, skiprows=[1])
    df.columns = [str(c).split(" (")[0].strip() for c in df.columns]
    df.index   = [str(i).split(" (")[0].strip() for i in df.index]
    df = df.dropna(how="all")
    # Keep only the criteria columns — ignore Route, Type etc.
    missing = [c for c in criteria_names if c not in df.columns]
    if missing:
        raise ValueError(f"Criteria columns not found in Alternatives_Data: {missing}")
    return df[criteria_names]

# ══════════════════════════════════════════════════════════════
# 3. AHP WEIGHT CALCULATION
# ══════════════════════════════════════════════════════════════
def ahp_criteria_weights(matrix):
    norm = matrix / matrix.sum(axis=0)
    weights = norm.mean(axis=1)
    n = matrix.shape[0]
    lmax = np.mean(np.dot(matrix.values, weights.values) / weights.values)
    CI = (lmax - n) / (n - 1)
    RI = {1:0,2:0,3:.58,4:.90,5:1.12,6:1.24,7:1.32,8:1.41,9:1.45,10:1.49}
    CR = CI / RI[n] if RI.get(n) else 0
    return weights, CI, CR

# ══════════════════════════════════════════════════════════════
# 4. READ BENEFIT/COST CONFIG
# ══════════════════════════════════════════════════════════════
def read_criteria_config(excel_file):
    cfg = pd.read_excel(excel_file, sheet_name="Criteria_Config", header=2)
    cfg.columns = cfg.columns.str.strip()
    cfg = cfg.dropna(subset=["Criteria Name"])
    cfg = cfg[cfg["Criteria Name"].astype(str).str.strip() != ""]
    cfg = cfg[~cfg["Criteria Name"].astype(str).str.startswith("⚠")]
    return cfg[cfg["Type (Benefit/Cost)"].str.strip().str.lower()=="benefit"]["Criteria Name"].tolist()

# ══════════════════════════════════════════════════════════════
# 5. NORMALIZE
# ══════════════════════════════════════════════════════════════
def normalize(df, benefit_criteria):
    out = pd.DataFrame(index=df.index)
    for col in df.columns:
        out[col] = df[col]/df[col].max() if col in benefit_criteria else df[col].min()/df[col]
    return out

# ══════════════════════════════════════════════════════════════
# 6. SAVE RESULTS TO EXCEL
# ══════════════════════════════════════════════════════════════
def save_results(excel_file, weights, CI, CR, norm_df, ranking, benefit_criteria):
    wb = load_workbook(excel_file)
    if "Results" in wb.sheetnames: del wb["Results"]
    ws = wb.create_sheet("Results")

    DARK="1F4E79"; MID="2E75B6"; LB="D9E2F3"; AB="EBF3FB"
    GOLD="FFF2CC"; WHITE="FFFFFF"; GREY="595959"; GRN="375623"; RED="C00000"

    def bdr():
        s=Side(style="thin"); return Border(left=s,right=s,top=s,bottom=s)
    def fill(c): return PatternFill("solid",start_color=c)
    def hf(col=WHITE,sz=11): return Font(name="Arial",bold=True,color=col,size=sz)
    def df_(col="000000",sz=10,bold=False): return Font(name="Arial",color=col,size=sz,bold=bold)

    def w(row,col,val,font=None,bg=None,align="center",fmt=None,brd=True):
        c=ws.cell(row=row,column=col,value=val)
        if font: c.font=font
        if bg:   c.fill=fill(bg)
        c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
        if fmt:  c.number_format=fmt
        if brd:  c.border=bdr()
        return c

    def sec(r,c1,c2,title):
        ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
        w(r,c1,title,font=hf(WHITE,12),bg=DARK,align="left"); ws.row_dimensions[r].height=22

    def hdr(r,cols,lbls):
        for col,lbl in zip(cols,lbls): w(r,col,lbl,font=hf(WHITE,10),bg=MID)
        ws.row_dimensions[r].height=18

    ws.merge_cells("A1:G1")
    w(1,1,"AHP–WSM RESULTS: BEST BUS ALLOCATION — TIRUPPUR REGION",
      font=Font(name="Arial",bold=True,size=14,color=WHITE),bg=DARK,align="center",brd=False)
    ws.row_dimensions[1].height=32
    ws.merge_cells("A2:G2")
    w(2,1,f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}  |  Data: Real TNSTC/SETC Routes",
      font=Font(name="Arial",italic=True,size=9,color=GREY),align="left",brd=False)
    ws.row_dimensions[2].height=16

    r=4
    sec(r,1,4,"  SECTION 1 — Criteria Weights (AHP)"); r+=1
    hdr(r,[1,2,3,4],["#","Criterion","Weight","Weight (%)"]); r+=1
    for i,(crit,wt) in enumerate(weights.items(),1):
        bg=LB if i%2 else AB
        w(r,1,i,font=df_(),bg=bg); w(r,2,crit,font=df_(),bg=bg,align="left")
        w(r,3,round(wt,6),font=df_(),bg=bg,fmt="0.000000")
        w(r,4,round(wt*100,2),font=df_(),bg=bg,fmt='0.00"%"'); r+=1
    w(r,1,"",font=df_(bold=True),bg=GOLD); w(r,2,"TOTAL",font=df_(bold=True),bg=GOLD,align="left")
    w(r,3,round(weights.sum(),4),font=df_(bold=True),bg=GOLD,fmt="0.0000")
    w(r,4,100.0,font=df_(bold=True),bg=GOLD,fmt='0.00"%"'); r+=2

    sec(r,1,4,"  SECTION 2 — Consistency Check"); r+=1
    hdr(r,[1,2,3,4],["Metric","Value","Threshold","Status"]); r+=1
    for lbl,val,ok in [("CI",CI,CI<.1),("CR",CR,CR<.1)]:
        sc=GRN if ok else RED
        w(r,1,f"Consistency {lbl} ({lbl})",font=df_(),bg=LB,align="left")
        w(r,2,round(val,4),font=df_(),bg=LB,fmt="0.0000"); w(r,3,"< 0.10",font=df_(),bg=LB)
        w(r,4,"✔  CONSISTENT" if ok else "✘  INCONSISTENT",
          font=Font(name="Arial",size=10,bold=True,color=sc),bg=LB); r+=1
    r+=1

    nc=len(norm_df.columns)
    sec(r,1,nc+2,"  SECTION 3 — Normalized Bus Performance Data (Real TNSTC Routes)"); r+=1
    hdr(r,list(range(1,nc+3)),["#","Bus"]+list(norm_df.columns)); r+=1
    for ci,col in enumerate(norm_df.columns,3):
        ws.cell(row=r-1,column=ci).value += " (B)" if col in benefit_criteria else " (C)"
    for i,(bus,rd) in enumerate(norm_df.iterrows(),1):
        bg=LB if i%2 else AB
        w(r,1,i,font=df_(),bg=bg); w(r,2,bus,font=df_(),bg=bg,align="left")
        for ci,val in enumerate(rd,3): w(r,ci,round(val,6),font=df_(),bg=bg,fmt="0.000000")
        r+=1
    r+=1

    sec(r,1,5,"  SECTION 4 — Best Bus Allocation Ranking (AHP–WSM) — Real TNSTC Data"); r+=1
    hdr(r,[1,2,3,4,5],["Rank","Bus","Route","Final Score","Recommendation"]); r+=1

    # Load route info for display
    route_info = {}
    try:
        df_alt = pd.read_excel(excel_file, sheet_name="Alternatives_Data",
                               index_col=0, header=0, skiprows=[1])
        for idx, row in df_alt.iterrows():
            route_info[str(idx)] = f"{row.get('Route','')}"
    except: pass

    medals={1:"🥇 BEST ALLOCATION",2:"🥈 2nd Choice",3:"🥉 3rd Choice"}
    for pos,(_,rd) in enumerate(ranking.iterrows(),1):
        bg,fc=(GOLD,DARK) if pos==1 else (LB if pos%2 else AB,"000000")
        route=route_info.get(rd["Bus"],"")
        w(r,1,pos,font=df_(fc,bold=(pos==1)),bg=bg)
        w(r,2,rd["Bus"],font=df_(fc,bold=(pos==1)),bg=bg,align="left")
        w(r,3,route,font=df_(fc,bold=(pos==1)),bg=bg,align="left")
        w(r,4,round(rd["Final Score"],6),font=df_(fc,bold=(pos==1)),bg=bg,fmt="0.000000")
        w(r,5,medals.get(pos,f"  {pos}th"),font=df_(fc,bold=(pos==1)),bg=bg); r+=1

    for i,width in enumerate([6,12,35,14,20],1):
        ws.column_dimensions[get_column_letter(i)].width=width
    ws.freeze_panes="A3"
    wb.save(excel_file)
    print(f"\n✔  Results saved → '{excel_file}' (sheet: Results)")

# ══════════════════════════════════════════════════════════════
# 7. LAUNCH EXPERT TOOL & WAIT
# ══════════════════════════════════════════════════════════════
def launch_expert_tool_and_wait():
    if not os.path.exists(EXPERT_TOOL):
        print("⚠  expert_tool.py not found — skipping."); return
    if os.path.exists(SENTINEL): os.remove(SENTINEL)

    print("\n" + "="*58)
    print("  STEP 1 — Expert Pairwise Comparison Input")
    print("="*58)
    print("  1. Browser opens automatically")
    print("  2. Fill in comparisons using the Saaty scale")
    print("  3. Click  'Save to Excel'  button")
    print("  4. This terminal continues automatically")
    print("="*58 + "\n")

    proc = subprocess.Popen(
        [sys.executable, EXPERT_TOOL,
         "--mode=subprocess", f"--sentinel={SENTINEL}"],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
    )
    time.sleep(1.5)
    webbrowser.open("http://127.0.0.1:5050")

    print("  Waiting for expert input", end="", flush=True)
    for elapsed in range(600):
        if os.path.exists(SENTINEL):
            os.remove(SENTINEL); print("  ✔  Saved!"); break
        time.sleep(1)
        if elapsed % 10 == 9: print(".", end="", flush=True)
    else:
        print("\n  ⚠  Timed out — using existing matrix.")

    try: proc.terminate(); proc.wait(timeout=3)
    except: pass

# ══════════════════════════════════════════════════════════════
# 8. MAIN
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    launch_expert_tool_and_wait()

    print("\n" + "="*58)
    print("  STEP 2 — Running AHP-WSM Analysis")
    print("="*58)

    criteria_matrix = read_criteria_matrix(EXCEL_FILE)
    if criteria_matrix.index.tolist() != criteria_matrix.columns.tolist():
        raise ValueError(f"Criteria row/col mismatch.\n  Rows={criteria_matrix.index.tolist()}\n  Cols={criteria_matrix.columns.tolist()}")

    weights, CI, CR = ahp_criteria_weights(criteria_matrix)
    print("\n=== CRITERIA WEIGHTS ===")
    print(weights.to_string())
    print(f"CI={CI:.4f},  CR={CR:.4f}")
    if CR >= 0.10:
        raise ValueError(f"Inconsistent judgments (CR={CR:.4f} ≥ 0.10). Re-run and revise.")

    criteria_names = weights.index.tolist()
    benefit_criteria = read_criteria_config(EXCEL_FILE)
    print(f"\nBenefit criteria: {benefit_criteria}")

    alternatives = read_alternatives(EXCEL_FILE, criteria_names)
    alternatives.index.name = "Bus"

    norm = normalize(alternatives, benefit_criteria)
    scores = norm.dot(weights.reindex(norm.columns))

    ranking = pd.DataFrame({
        "Bus": scores.index, "Final Score": scores.values
    }).sort_values("Final Score", ascending=False).reset_index(drop=True)
    ranking.index += 1

    print("\n=== TOP 10 BEST BUS ALLOCATIONS (Real TNSTC Data) ===")
    print(ranking.head(10).to_string())
    print("\n=== BOTTOM 5 ===")
    print(ranking.tail(5).to_string())

    save_results(EXCEL_FILE, weights.reindex(norm.columns), CI, CR,
                 norm, ranking, benefit_criteria)
